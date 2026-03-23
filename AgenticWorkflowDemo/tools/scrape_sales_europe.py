"""
Scrape DailyRemote sales job listings across multiple European countries.
Outputs deduplicated results to an Excel file.

Usage (from AgenticWorkflowDemo/ directory):
    python tools/scrape_sales_europe.py
    python tools/scrape_sales_europe.py --output sales_europe.xlsx --max-pages 50
"""

import argparse
import time
import sys
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE_URL = "https://dailyremote.com"
DEFAULT_SEARCH = "sales"
DEFAULT_OUTPUT = "sales_europe.xlsx"
REQUEST_DELAY = 1.5  # seconds between requests

EUROPEAN_COUNTRIES = [
    "United Kingdom",
    "Germany",
    "France",
    "Netherlands",
    "Spain",
    "Italy",
    "Poland",
    "Sweden",
    "Belgium",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def build_url(search: str, location: str, page: int) -> str:
    from urllib.parse import urlencode
    params = {
        "search": search,
        "location_country": location,
        "page": page,
    }
    return f"{BASE_URL}/?{urlencode(params)}"


def extract_job_data(card) -> dict:
    """Extract structured data from a single job card <article> element."""
    job = {
        "title": "",
        "company": "N/A (Premium)",
        "description": "",
        "location": "",
        "date_posted": "",
        "job_url": "",
    }

    # --- Title + URL ---
    title_h2 = card.find("h2", class_="job-position")
    if title_h2:
        link = title_h2.find("a", href=True)
        if link:
            href = link.get("href", "")
            job["job_url"] = href if href.startswith("http") else BASE_URL + href
            job["title"] = link.get_text(strip=True)

    # --- Company + Date Posted ---
    company_div = card.find("div", class_="company-name")
    if company_div:
        spans = company_div.find_all("span", recursive=False)
        # If more than 3 spans, the first may be a real company name (paywall unlocked)
        if len(spans) > 3:
            job["company"] = spans[0].get_text(strip=True)
        # Last non-separator span = date posted
        date_spans = [s for s in spans if s.get_text(strip=True) not in ("·", "")]
        if len(date_spans) >= 2:
            job["date_posted"] = date_spans[-1].get_text(strip=True)

    # --- Location ---
    job_meta = card.find("div", class_="job-meta")
    if job_meta:
        for tag_span in job_meta.find_all("span", class_="card-tag"):
            text = tag_span.get_text(strip=True)
            if "\U0001f30e" in text or "\U0001f30d" in text or "\U0001f30f" in text:
                location_text = (
                    text.replace("\U0001f30e", "")
                        .replace("\U0001f30d", "")
                        .replace("\U0001f30f", "")
                        .strip()
                )
                job["location"] = location_text
                break

    # --- Description (AI-generated responsibilities blurb) ---
    desc_div = card.find("div", class_="ai-responsibilities")
    if desc_div:
        job["description"] = desc_div.get_text(strip=True)

    return job


def parse_job_cards(soup: BeautifulSoup) -> list[dict]:
    jobs = []
    cards = soup.find_all("article", class_="card")
    for card in cards:
        job = extract_job_data(card)
        if job.get("title"):
            jobs.append(job)
    return jobs


def scrape_country(search: str, country: str, max_pages: int, session: requests.Session) -> list[dict]:
    all_jobs = []
    print(f"\n[{country}] Starting scrape...")

    for page in range(1, max_pages + 1):
        url = build_url(search, country, page)
        print(f"  Page {page}: {url}")

        try:
            resp = session.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"  ERROR: {e}", file=sys.stderr)
            break

        soup = BeautifulSoup(resp.text, "lxml")
        jobs = parse_job_cards(soup)
        print(f"  Found {len(jobs)} jobs")

        if not jobs:
            print(f"  No jobs on page {page} — stopping.")
            break

        all_jobs.extend(jobs)

        if page < max_pages:
            time.sleep(REQUEST_DELAY)

    print(f"  [{country}] Total: {len(all_jobs)} jobs")
    return all_jobs


def deduplicate(jobs: list[dict]) -> list[dict]:
    seen = set()
    unique = []
    for job in jobs:
        key = job.get("job_url", "")
        if key and key not in seen:
            seen.add(key)
            unique.append(job)
        elif not key:
            unique.append(job)
    return unique


def save_to_excel(jobs: list[dict], output_path: str) -> None:
    columns = ["title", "company", "description", "location", "date_posted", "job_url"]
    headers = ["Title", "Company", "Description", "Location", "Date Posted", "Job URL"]

    df = pd.DataFrame(jobs, columns=columns)
    df.columns = headers

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sales Jobs Europe")

        ws = writer.sheets["Sales Jobs Europe"]

        # Column width caps
        col_widths = {
            "A": 45,  # Title
            "B": 30,  # Company
            "C": 80,  # Description
            "D": 25,  # Location
            "E": 18,  # Date Posted
            "F": 60,  # Job URL
        }

        for col_letter, max_w in col_widths.items():
            col_idx = ord(col_letter) - ord("A")
            col_cells = [ws.cell(row=r, column=col_idx + 1) for r in range(1, ws.max_row + 1)]
            content_w = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_letter].width = min(content_w + 4, max_w)

        # Wrap text in Description column
        from openpyxl.styles import Alignment
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Freeze header row
        ws.freeze_panes = "A2"

    print(f"\nSaved {len(jobs)} jobs to: {output.resolve()}")


def main():
    parser = argparse.ArgumentParser(description="Scrape DailyRemote sales jobs across European countries")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel filename")
    parser.add_argument("--max-pages", type=int, default=50, help="Max pages per country")
    args = parser.parse_args()

    print(f"Scraping DailyRemote: search='sales', countries={EUROPEAN_COUNTRIES}")
    print(f"Output: {args.output}\n")

    session = requests.Session()
    all_jobs = []

    for country in EUROPEAN_COUNTRIES:
        jobs = scrape_country(DEFAULT_SEARCH, country, args.max_pages, session)
        all_jobs.extend(jobs)

    print(f"\nRaw total: {len(all_jobs)} jobs")
    unique_jobs = deduplicate(all_jobs)
    print(f"After deduplication: {len(unique_jobs)} unique jobs")

    if not unique_jobs:
        print("No jobs found. Check selectors or network connectivity.")
        sys.exit(1)

    save_to_excel(unique_jobs, args.output)


if __name__ == "__main__":
    main()
